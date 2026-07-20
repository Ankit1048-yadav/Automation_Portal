import pandas as pd
import msoffcrypto
import io
from pyxlsb import open_workbook
from openpyxl.styles import Font, Border, Side

    # -------- CONFIG --------
CR_TO_NORMAL = 10000000

def generate_report(FILE_PATH, PASSWORD, OUTPUT_FILE):


    # -------- READ FILE --------
    decrypted = io.BytesIO()

    with open(FILE_PATH, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=PASSWORD)
        office.decrypt(decrypted)

    decrypted.seek(0)

    with open_workbook(decrypted) as wb:
        sheet = wb.get_sheet(wb.sheets[0])
        rows = [[item.v for item in row] for row in sheet.rows()]

    df = pd.DataFrame(rows[1:], columns=rows[0])
    df.columns = df.columns.str.strip().str.upper()

    # -------- CLEAN DATA --------
    df["POS"] = pd.to_numeric(df["POS"], errors="coerce").fillna(0)
    df["AMOUNT COLLECTED"] = pd.to_numeric(df["AMOUNT COLLECTED"], errors="coerce").fillna(0)
    df["REPORTING"] = pd.to_numeric(df["REPORTING"], errors="coerce").fillna(0)

    df["UPGRADE/RECOVERY"] = df["UPGRADE/RECOVERY"].astype(str).str.upper()
    df["RSL STATUS"] = df["RSL STATUS"].astype(str).str.upper()
    df["RESOLUTION"] = df["RESOLUTION"].astype(str).str.upper()
    df["ZONE"] = df["ZONE"].astype(str).str.upper().str.strip()

    df["BKT"] = df["BKT"].astype(str)
    df["BKT"] = df["BKT"].apply(lambda x: "B4" if x.startswith("B4") else x)

    # -------- POS BRACKET CREATION --------
    def pos_bracket(pos):
        if pos < 1000:
            return "0k-1k"
        elif 1000 <= pos <= 5000:
            return "1K-5K"
        elif 5000 <= pos <= 10000:
            return "5K-10K"
        else:
            return "More than 10K"

    df["POS BRACKET"] = df["POS"].apply(pos_bracket)

    # -------- CALC FUNCTION --------
    def calc_summary(g):

        coa = g["ACCREF"].nunique()
        pos = g["POS"].sum()
        collection = g["AMOUNT COLLECTED"].sum()

        upgrade = g.loc[g["UPGRADE/RECOVERY"] == "UPGRADE", "POS"].sum()
        recovery_cr = g.loc[g["UPGRADE/RECOVERY"] == "RECOVERY", "REPORTING"].sum()
        recovery_normal = recovery_cr * CR_TO_NORMAL

        resolved = g.loc[g["RSL STATUS"] == "RESOLVED", "POS"].sum()

        norm = g.loc[g["RESOLUTION"] == "NR", "POS"].sum()
        rb = g.loc[g["RESOLUTION"] == "RB", "POS"].sum()
        st = g.loc[g["RESOLUTION"] == "ST", "POS"].sum()

        ror = (collection / pos * 100) if pos else 0
        upgrade_pct = (upgrade / pos * 100) if pos else 0
        resolution_pct = (resolved / pos * 100) if pos else 0
        norm_pct = (norm / pos * 100) if pos else 0

        ur_cr = round((upgrade + recovery_normal) / CR_TO_NORMAL, 2)

        return [coa,pos,collection,ror,upgrade,upgrade_pct,recovery_normal,
                resolved,resolution_pct,norm,rb,st,norm_pct,ur_cr]

    # -------- COMMON COLUMNS --------
    cols = ["COA","POS","Collection","ROR %","Upgrade","Upgrade %",
            "Recovery","Resolution","Resolution %","Norm","RB","ST",
            "Norm %","U+R (CR)"]

    percent_cols = ["ROR %","Upgrade %","Resolution %","Norm %"]

    # -------- BKT SUMMARY --------
    bkt_rows=[]
    for b,g in df.groupby("BKT"):
        bkt_rows.append([b]+calc_summary(g))

    bkt_df = pd.DataFrame(bkt_rows, columns=["BKT"]+cols)
    bkt_df.loc[len(bkt_df)] = ["Grand Total"] + calc_summary(df)
    bkt_df[percent_cols] /= 100

    # -------- ZONE SUMMARY --------
    zone_rows=[]
    for z,g in df.groupby("ZONE"):
        zone_rows.append([z]+calc_summary(g))

    zone_df = pd.DataFrame(zone_rows, columns=["ZONE"]+cols)
    zone_df.loc[len(zone_df)] = ["Grand Total"] + calc_summary(df)
    zone_df[percent_cols] /= 100

    # -------- POS BRACKET SUMMARY --------
    pos_rows=[]
    for p,g in df.groupby("POS BRACKET"):
        pos_rows.append([p]+calc_summary(g))

    order = ["0k-1K","1K-5K","5K-10K","More than 10K"]
    pos_df = pd.DataFrame(pos_rows, columns=["POS BRACKET"]+cols)
    pos_df["POS BRACKET"] = pd.Categorical(pos_df["POS BRACKET"], categories=order, ordered=True)
    pos_df = pos_df.sort_values("POS BRACKET")

    pos_df.loc[len(pos_df)] = ["Grand Total"] + calc_summary(df)
    pos_df[percent_cols] /= 100

    # -------- WRITE TO EXCEL --------
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

        bkt_df.to_excel(writer, sheet_name="Summary", index=False, startrow=1)
        ws = writer.book["Summary"]
        ws.cell(1,1,"BKT SUMMARY").font = Font(bold=True)

        zone_start = len(bkt_df) + 4
        zone_df.to_excel(writer, sheet_name="Summary", index=False, startrow=zone_start)
        ws.cell(zone_start,1,"ZONE SUMMARY").font = Font(bold=True)

        pos_start = zone_start + len(zone_df) + 4
        pos_df.to_excel(writer, sheet_name="Summary", index=False, startrow=pos_start)
        ws.cell(pos_start,1,"POS BRACKET SUMMARY").font = Font(bold=True)

        # -------- FORMATTING --------
        thin = Border(left=Side(style='thin'),right=Side(style='thin'),
                    top=Side(style='thin'),bottom=Side(style='thin'))

        indian_format = '#,##,##0'
        indian_format_2 = '#,##,##0.00'

        header_row = 2
        header_map = {}
        for c in range(1, ws.max_column+1):
            header_map[ws.cell(header_row,c).value] = c

        percent_excel_cols = [header_map[h] for h in percent_cols if h in header_map]

        for r in range(1, ws.max_row+1):
            for c in range(1, ws.max_column+1):

                cell = ws.cell(r,c)

                # Apply border ONLY if cell has value
                if cell.value not in (None, ""):
                    cell.border = thin

                if isinstance(cell.value,(int,float)):

                    if c in percent_excel_cols:
                        cell.number_format="0.00%"

                    elif ws.cell(header_row,c).value=="U+R (CR)":
                        cell.number_format=indian_format_2

                    else:
                        cell.number_format=indian_format

        # -------- GRAND TOTAL BOLD --------
        for r in range(1, ws.max_row+1):
            if ws.cell(r,1).value=="Grand Total":
                for c in range(1, ws.max_column+1):
                    ws.cell(r,c).font = Font(bold=True)

    return OUTPUT_FILE
