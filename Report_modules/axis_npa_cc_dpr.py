import pandas as pd
import msoffcrypto
import io
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


def generate_report(input_file, password, output_file):

    decrypted = io.BytesIO()

    with open(input_file, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)

    decrypted.seek(0)

    # READ SHEET
    df = pd.read_excel(
        decrypted,
        engine="pyxlsb",
        sheet_name=0,
        dtype=str
    )

    account_col = "ACCREF"
    total_col = "Amount Collected"

    df[account_col] = df[account_col].astype(str).str.strip()
    df[total_col] = pd.to_numeric(df[total_col], errors="coerce")

    date_cols = []
    date_map = {}

    for col in df.columns:

        if str(col).isdigit():

            real_date = datetime(1899, 12, 30) + timedelta(days=int(col))

            date_cols.append(col)
            date_map[col] = real_date

    date_cols = sorted(date_cols, key=lambda x: int(x))

    df = df[df[total_col] > 0].copy()

    def get_first_payment_date(row):

        for col in date_cols:

            try:

                value = float(row[col])

                if value > 0:
                    return date_map[col].date()

            except:
                pass

        return None

    df["Date"] = df.apply(get_first_payment_date, axis=1)

    final_df = pd.DataFrame({
        "Account no": df[account_col],
        "Collection": df[total_col],
        "Date": df["Date"],
        "Team-Digital": ""
    })

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    final_df.to_excel(
        output_file,
        index=False
    )

    wb = load_workbook(output_file)
    ws = wb.active

    max_row = ws.max_row

    ws["E1"] = "Lien check"
    ws["F1"] = "Lien"
    ws["G1"] = "Net Collection"

    for row in range(2, max_row + 1):

        ws[f"E{row}"] = f'=IF(D{row}="Digital",INT(B{row})=B{row},"")'
        ws[f"F{row}"] = f'=IF(E{row}=FALSE,B{row},0)'
        ws[f"G{row}"] = f'=B{row}-F{row}'

    header_fill = PatternFill(
        start_color="C00000",
        end_color="C00000",
        fill_type="solid"
    )

    header_font = Font(
        size=9,
        bold=True,
        color="FFFFFF"
    )

    data_font = Font(size=9)

    center_align = Alignment(
        horizontal="center",
        vertical="center"
    )

    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in ws.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=7
    ):
        for cell in row:

            cell.border = border
            cell.alignment = center_align
            cell.font = data_font

    for col in range(1, 8):

        cell = ws.cell(row=1, column=col)

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for row in ws.iter_rows(
        min_row=2,
        min_col=1,
        max_col=1
    ):
        for cell in row:
            cell.number_format = "@"

    for row in ws.iter_rows(
        min_row=2,
        min_col=3,
        max_col=3
    ):
        for cell in row:
            cell.number_format = "DD-MM-YYYY"

    wb.save(output_file)

    return output_file