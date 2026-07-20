import pandas as pd
import numpy as np
import datetime
import io
import msoffcrypto
import os
import re

from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter


def generate_report(files, password, month, year, output_file):
    
    print("Password Received:", password)

    current_month = int(month)
    current_year = int(year)

    DATE_COLUMN = "POSTING_DT"
    ACCOUNT_COLUMN = "ACC_NUM"
    ACCOUNT_LENGTH = 19

# =====================================
# READ PASSWORD PROTECTED EXCEL
# =====================================
    def read_protected_excel(file_path, password):

        decrypted = io.BytesIO()
        try:

            with open(file_path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)
                office_file.decrypt(decrypted)
        except Exception as e:
            print(f"Error reading {file}: {e}")
            raise e

        decrypted.seek(0)

        return pd.read_excel(
            decrypted,
            engine="openpyxl",
            dtype={"ACC_NUM": str}
        )


    # =====================================
    # ALLOCATION EXTRACT
    # =====================================

    def extract_allocation_from_filename(file_name):

        month_pairs = [
            "Jan-Feb", "Feb-Mar", "Mar-Apr", "Apr-May",
            "May-June", "June-July", "July-Aug", "Aug-Sep",
            "Sep-Oct", "Oct-Nov", "Nov-Dec", "Dec-Jan"
        ]

        for m in month_pairs:
            if m in file_name:
                return m

        return ""


    # =====================================
    # EXTRACT CYCLE
    # =====================================

    def extract_cycle(product_name):

        match = re.search(r"(Cycle\s*\d+)", product_name, re.IGNORECASE)

        if match:
            return match.group(1).strip()

        return ""


    # =====================================
    # FORMAT WORKSHEET
    # =====================================

    def format_worksheet(ws):

        thin = Side(style="thin")

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        header_fill = PatternFill(
            start_color="D9E1F2",
            end_color="D9E1F2",
            fill_type="solid"
        )

        for row in ws.iter_rows():
            for cell in row:

                cell.border = border

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

        # Header Style
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Auto column width
        for col in ws.columns:

            max_length = 0

            col_letter = get_column_letter(col[0].column)

            for cell in col:

                try:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2


    # =====================================
    # REQUIRED COLUMNS
    # =====================================

    required_columns = [
        "ACC_NUM",
        "VALUE_DATE",
        "POSTING_DT",
        "TXN_AMOUNT",
        "DESCRIPTION",
        "BKT",
        "AGENCY",
        "REMARK",
        "BEFOR ALLOCATION"
    ]


    # =====================================
    # MAIN PROCESS
    # =====================================

    all_files = files

    df_list = []

    for file in all_files:

        if os.path.basename(file).startswith("~$"):
            continue

        try:

            print(f"Processing: {os.path.basename(file)}")

            df = read_protected_excel(file, password)

            # =====================================
            # CREATE MISSING COLUMNS
            # =====================================

            for col in required_columns:

                if col not in df.columns:

                    if col in ["TXN_AMOUNT"]:
                        df[col] = 0.0
                    else:
                        df[col] = ""

                    print(f"Missing column created: {col}")

            # =====================================
            # ACCOUNT FORMAT
            # =====================================

            df[ACCOUNT_COLUMN] = (
                df[ACCOUNT_COLUMN]
                .astype(str)
                .str.replace(".0", "", regex=False)
                .str.zfill(ACCOUNT_LENGTH)
            )

            # =====================================
            # DATE CONVERSION
            # =====================================

            df["VALUE_DATE"] = pd.to_datetime(
                df["VALUE_DATE"],
                errors='coerce',
                dayfirst=True
            )

            df["POSTING_DT"] = pd.to_datetime(
                df["POSTING_DT"],
                errors='coerce',
                dayfirst=True
            )

            # =====================================
            # FILTER MONTH & YEAR
            # =====================================

            df = df[
                (df["POSTING_DT"].dt.month == current_month) &
                (df["POSTING_DT"].dt.year == current_year)
            ]

            # =====================================
            # FILE INFO
            # =====================================

            file_name_only = os.path.splitext(
                os.path.basename(file)
            )[0]

            df["Reference"] = file_name_only

            df["Allocation"] = extract_allocation_from_filename(
                file_name_only
            )

            # =====================================
            # CLEAN DATA
            # =====================================

            df["TXN_AMOUNT"] = pd.to_numeric(
                df["TXN_AMOUNT"],
                errors="coerce"
            ).fillna(0.0)

            df["DESCRIPTION"] = (
                df["DESCRIPTION"]
                .astype(str)
                .str.upper()
                .str.strip()
            )

            df["REMARK"] = (
                df["REMARK"]
                .astype(str)
                .str.upper()
                .str.strip()
            )

            df["BKT"] = (
                df["BKT"]
                .astype(str)
                .str.upper()
                .str.strip()
            )

            # =====================================
            # LIEN
            # =====================================

            df["Lien"] = np.where(
                (
                    (df["DESCRIPTION"] == "AUTO DEBIT - CA - PAYMENT RECD") |
                    (df["REMARK"] == "DO NOT CONSIDER FOR BILLING PURPOSE") |
                    (df["REMARK"] == "NOT CONSIDER")
                ),
                df["TXN_AMOUNT"],
                0.0
            )

            # =====================================
            # NET
            # =====================================

            df["Net"] = df["TXN_AMOUNT"] - df["Lien"]

            # =====================================
            # PRODUCT
            # =====================================

            df["Product"] = np.where(
                df["BKT"].str.contains("WOF", na=False),
                "WO",
                df["Reference"].apply(extract_cycle)
            )

            # =====================================
            # FINAL COLUMNS
            # =====================================

            final_columns = [
                "ACC_NUM",
                "VALUE_DATE",
                "POSTING_DT",
                "TXN_AMOUNT",
                "Lien",
                "Net",
                "DESCRIPTION",
                "BKT",
                "AGENCY",
                "Product",
                "Allocation",
                "Reference",
                "REMARK",
                "BEFOR ALLOCATION"
            ]

            df = df[final_columns]

            df_list.append(df)

        except Exception as e:

            print(f"Error reading {file}: {e}")


    # =====================================
    # FINAL OUTPUT
    # =====================================
    print("DF LIST LENGTH =", len(df_list))
    if not df_list:
        raise ValueError("WRONG_PASSWORD")

    

    final_df = pd.concat(df_list, ignore_index=True)

    # =====================================
    # SUMMARY SHEET
    # =====================================

    summary_df = final_df.groupby(
        "ACC_NUM",
        as_index=False
    ).agg({
        "Product": "min",
        "TXN_AMOUNT": "sum",
        "POSTING_DT": "min",
        "Lien": "sum",
        "Net": "sum"
    })

    summary_df.rename(columns={
        "ACC_NUM": "Account No",
        "TXN_AMOUNT": "Collection",
        "POSTING_DT": "Date",
        "Net": "Net Collection"
    }, inplace=True)

    summary_df["Team-Digital"] = ""

    summary_df = summary_df[
        [
            "Account No",
            "Product",
            "Collection",
            "Date",
            "Lien",
            "Net Collection",
            "Team-Digital"
        ]
    ]
    print("STARTING EXCEL WRITE")
    # =====================================
    # WRITE EXCEL
    # =====================================

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        final_df.to_excel(
            writer,
            sheet_name="Raw_Data",
            index=False
        )

        summary_df.to_excel(
            writer,
            sheet_name="Team-Digital",
            index=False
        )

        ws_raw = writer.sheets["Raw_Data"]
        ws_summary = writer.sheets["Team-Digital"]

        format_worksheet(ws_raw)
        format_worksheet(ws_summary)

        # Date format
        for ws in [ws_raw, ws_summary]:

            for row in ws.iter_rows(min_row=2):

                for cell in row:

                    if isinstance(cell.value, datetime.datetime):
                        cell.number_format = "DD-MM-YYYY"

        # Indian Currency Format
        indian_currency_format = '#,##,##0.00'

        raw_amount_cols = [
            "TXN_AMOUNT",
            "Lien",
            "Net"
        ]

        summary_amount_cols = [
            "Collection",
            "Lien",
            "Net Collection"
        ]

        for ws, cols in [
            (ws_raw, raw_amount_cols),
            (ws_summary, summary_amount_cols)
        ]:

            for col in ws[1]:

                if col.value in cols:

                    col_letter = col.column_letter

                    for cell in ws[col_letter][1:]:
                        cell.number_format = indian_currency_format
    print("EXCEL CREATED SUCCESSFULLY")
    return output_file