import pandas as pd
import re
from openpyxl.utils import get_column_letter

def generate_cmd_clean_report(input_file, output_file):

    df = pd.read_excel(input_file, dtype=str)

    df.columns = df.columns.astype(str)
    account_col = df.columns[0]
    df[account_col] = df[account_col].astype(str)

    email_col = None
    for col in df.columns:
        if df[col].astype(str).str.contains(r'@', na=False).any():
            email_col = col
            break

    def convert_to_general(x):
        if pd.isna(x):
            return ''
        x = str(x).strip().replace(' ', '')
        if x.startswith("'"):
            x = x[1:]
        return x

    cols_to_clean = [col for col in df.columns if col != account_col and col != email_col]

    for col in cols_to_clean:
        df[col] = df[col].apply(convert_to_general)

    def clean_number(x):
        if pd.isna(x):
            return ''
        x = str(x).strip().replace(' ', '')
        x = re.sub(r'[^\d]', '', x)
        x = re.sub(r'^(0{0,3}91)+', '', x)
        if len(x) > 10:
            x = x[-10:]
        if len(x) == 10 and x[0] in '6789':
            return x
        return ''

    for col in cols_to_clean:
        df[col] = df[col].apply(clean_number)

    for col in cols_to_clean:
        df[col] = df[col].apply(convert_to_general)

    melt_cols = [col for col in df.columns if col != account_col]
    df_unpivot = df.melt(id_vars=account_col, value_vars=melt_cols,
                         var_name='Attribute', value_name='Value')

    df_unpivot = df_unpivot[df_unpivot['Value'] != '']
    df_unpivot = df_unpivot.drop_duplicates(subset=[account_col, 'Attribute'])

    df_final = df_unpivot.pivot(index=account_col,
                                columns='Attribute',
                                values='Value').reset_index()

    def remove_duplicates_rowwise(row):
        seen = set()
        return [val if not (val in seen or seen.add(val)) else '' for val in row]

    cols_except_first = df_final.columns[1:]
    df_final[cols_except_first] = df_final[cols_except_first].apply(
        remove_duplicates_rowwise,
        axis=1,
        result_type='broadcast'
    )

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='CleanedData')

        wb = writer.book
        ws = writer.sheets['CleanedData']

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                 min_col=1, max_col=1):
            for cell in row:
                cell.number_format = '@'

        for col_idx in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = 'General'