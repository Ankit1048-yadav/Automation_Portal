import pandas as pd
import msoffcrypto
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


# ==============================
# MAIN FUNCTION
# ==============================
def generate_report(files, password, month, year, output_file):

    TARGET_MONTH = int(month)
    TARGET_YEAR = int(year)

    # ==============================
    # READ PASSWORD PROTECTED XLSB
    # ==============================
    def read_protected_xlsb(file_path, password):

        decrypted = BytesIO()

        with open(file_path, "rb") as f:
            office = msoffcrypto.OfficeFile(f)
            office.load_key(password=password)
            office.decrypt(decrypted)

        decrypted.seek(0)

        df = pd.read_excel(decrypted, engine="pyxlsb")
        df.columns = df.columns.map(lambda x: str(x).strip())

        return df

    # ==============================
    # PROCESS FILE
    # ==============================
    def process_file(file_path, password):

        df = read_protected_xlsb(file_path, password)

        if df.empty or "ACCOUNT_NO" not in df.columns or "FINAL_ZONE" not in df.columns:
            return pd.DataFrame()

        df["ACCOUNT_NO"] = (
            df["ACCOUNT_NO"]
            .astype(str)
            .str.split(".")
            .str[0]
            .str.strip()
            .str.zfill(15)
        )

        def tag(zone):
            z = str(zone).strip().lower().replace(" ", "").replace("-", "")
            return "Bangalore" if z in ["south", "south1", "south2"] else "Noida"

        df["Central_Tagging"] = df["FINAL_ZONE"].apply(tag)

        # ==============================
        # DATE COLUMNS DETECTION
        # ==============================
        serial_date_map = {}

        for col in df.columns:
            try:
                num = float(col)
                serial_date_map[col] = pd.to_datetime(
                    num,
                    origin="1899-12-30",
                    unit="D"
                )
            except:
                continue

        if not serial_date_map:
            return pd.DataFrame()

        today = pd.Timestamp.today()

        valid_dates = {
            c: d for c, d in serial_date_map.items()
            if d <= today
        }

        date_cols = [
            c for c, d in valid_dates.items()
            if d.month == TARGET_MONTH and d.year == TARGET_YEAR
        ]

        if not date_cols:
            print(f"⚠️ No data found for {TARGET_MONTH}-{TARGET_YEAR}")
            return pd.DataFrame()

        df[date_cols] = (
            df[date_cols]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0)
        )

        date_cols = sorted(date_cols, key=lambda x: valid_dates[x])

        df["Total_Collection"] = df[date_cols].sum(axis=1)

        def first_date(row):
            for c in date_cols:
                if row[c] > 0:
                    return valid_dates[c]
            return None

        df["First_Collection_Date"] = df.apply(first_date, axis=1)

        df = df[df["Total_Collection"] > 0]

        df["File_Source"] = os.path.basename(file_path)

        return df[
            [
                "ACCOUNT_NO",
                "Central_Tagging",
                "Total_Collection",
                "First_Collection_Date",
                "File_Source"
            ]
        ]

    # ==============================
    # RUN FOR ALL FILES
    # ==============================
    dfs = []

    for f in files:
        temp = process_file(f, password)
        if not temp.empty:
            dfs.append(temp)

    if dfs:
        final_df = pd.concat(dfs, ignore_index=True)
    else:
        final_df = pd.DataFrame(columns=[
            "ACCOUNT_NO",
            "Central_Tagging",
            "Total_Collection",
            "First_Collection_Date",
            "File_Source"
        ])

    # ==============================
    # TEAM SUMMARY
    # ==============================
    team_df = final_df.groupby(
        ["ACCOUNT_NO", "Central_Tagging"],
        as_index=False
    ).agg({
        "Total_Collection": "sum",
        "First_Collection_Date": "min"
    })

    team_df.rename(columns={
        "ACCOUNT_NO": "Account no",
        "Total_Collection": "Total Collection",
        "First_Collection_Date": "Date"
    }, inplace=True)

    team_df["Team-Digital"] = ""
    team_df["Lien check"] = ""
    team_df["Lien"] = ""
    team_df["Net Collection"] = ""

    # ==============================
    # SAVE FILE
    # ==============================
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        final_df.to_excel(writer, index=False, sheet_name="Raw_Data")
        team_df.to_excel(writer, index=False, sheet_name="Team-Digital")

        for sheet in ["Raw_Data", "Team-Digital"]:

            ws = writer.sheets[sheet]

            max_row = ws.max_row
            max_col = ws.max_column

            header_fill = PatternFill("solid", fgColor="C00000")
            header_font = Font(bold=True, color="FFFFFF")
            align = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # HEADER STYLE
            for c in range(1, max_col + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align
                cell.border = border

            # DATA STYLE
            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.alignment = align
                    cell.border = border

            # DATE FORMAT
            for c in range(1, max_col + 1):
                header = ws.cell(1, c).value
                if header in ["First_Collection_Date", "Date"]:
                    for r in range(2, max_row + 1):
                        ws.cell(r, c).number_format = "DD-MM-YYYY"

            # AUTO WIDTH
            for col in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = length + 2

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # TEAM FORMULAS
        ws = writer.sheets["Team-Digital"]

        for r in range(2, ws.max_row + 1):
            ws[f"F{r}"] = f'=IF(E{r}="Digital",INT(C{r})=C{r},"")'
            ws[f"G{r}"] = f'=IF(F{r}=FALSE,C{r},0)'
            ws[f"H{r}"] = f'=C{r}-G{r}'

    print("\n✅ File Created Successfully")
    print(f"📁 Output File: {output_file}")

    return output_file