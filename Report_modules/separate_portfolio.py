import pandas as pd
import os
import zipfile
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


def generate_separate_portfolio(input_file, output_file,separate_files=False):

    # =====================
    # READ FILE
    # =====================
    df = pd.read_excel(input_file)

    portfolio_column = "Portfolio"

# ==================================
# SEPARATE EXCEL FILES MODE
# ==================================
    if separate_files:

        temp_dir = tempfile.mkdtemp()

        created_files = []

        for portfolio, data in df.groupby(portfolio_column):

            safe_name = str(portfolio).replace("/", "-")

            file_path = os.path.join(
                temp_dir,
                f"{safe_name}.xlsx"
            )

            with pd.ExcelWriter(
                file_path,
                engine="openpyxl"
            ) as writer:

                data.to_excel(
                    writer,
                    sheet_name="Data",
                    index=False
                )

            created_files.append(file_path)

        with zipfile.ZipFile(
            output_file,
            "w",
            zipfile.ZIP_DEFLATED
        ) as zipf:

            for f in created_files:

                zipf.write(
                    f,
                    os.path.basename(f)
                )

        return output_file

    # =====================
    # CREATE SHEETS
    # =====================
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        for portfolio, data in df.groupby(portfolio_column):

            sheet_name = str(portfolio)[:31]

            data.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

    # =====================
    # LOAD WORKBOOK
    # =====================
    wb = load_workbook(output_file)

    # =====================
    # STYLES
    # =====================
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # =====================
    # FORMAT EACH SHEET
    # =====================
    for ws in wb.worksheets:

        max_row = ws.max_row
        max_col = ws.max_column

        headers = {}

        for col in range(1, max_col + 1):
            headers[
                ws.cell(row=1, column=col).value
            ] = col

        # =====================
        # X = V + W
        # =====================
        if max_col >= 24:

            for row in range(2, max_row + 1):
                ws[f"X{row}"] = f"=V{row}+W{row}"

        # =====================
        # VARIANCE FORMULA
        # =====================
        if (
            "Variance" in headers
            and "Gross" in headers
            and "AmountPaid" in headers
        ):

            variance_col = get_column_letter(
                headers["Variance"]
            )

            gross_col = get_column_letter(
                headers["Gross"]
            )

            amountpaid_col = get_column_letter(
                headers["AmountPaid"]
            )

            for row in range(2, max_row + 1):

                ws[f"{variance_col}{row}"] = (
                    f"={gross_col}{row}-{amountpaid_col}{row}"
                )

        # =====================
        # HEADER FORMAT
        # =====================
        for cell in ws[1]:

            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        # =====================
        # FREEZE HEADER
        # =====================
        ws.freeze_panes = "A2"

        # =====================
        # BORDERS
        # =====================
        for row in ws.iter_rows():

            for cell in row:

                cell.border = thin_border

        # =====================
        # AUTO WIDTH
        # =====================
        for column in ws.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except:
                    pass

            ws.column_dimensions[
                column_letter
            ].width = max_length + 3

        # =====================
        # CONDITIONAL FORMAT
        # =====================
        if "Variance" in headers:

            variance_col_letter = get_column_letter(
                headers["Variance"]
            )

            variance_range = (
                f"{variance_col_letter}2:"
                f"{variance_col_letter}{max_row}"
            )

            red_fill = PatternFill(
                start_color="FFC7CE",
                end_color="FFC7CE",
                fill_type="solid"
            )

            green_fill = PatternFill(
                start_color="C6EFCE",
                end_color="C6EFCE",
                fill_type="solid"
            )

            ws.conditional_formatting.add(
                variance_range,
                CellIsRule(
                    operator="lessThan",
                    formula=["0"],
                    fill=red_fill
                )
            )

            ws.conditional_formatting.add(
                variance_range,
                CellIsRule(
                    operator="greaterThan",
                    formula=["0"],
                    fill=green_fill
                )
            )

        # =====================
        # DATE FORMAT
        # =====================
        for col in ws.columns:

            header = str(
                ws.cell(
                    row=1,
                    column=col[0].column
                ).value
            ).lower()

            if "date" in header:

                for cell in col:

                    cell.number_format = "DD-MMM-YYYY"

    # =====================
    # SAVE FILE
    # =====================
    wb.save(output_file)
    wb.close()

    return output_file